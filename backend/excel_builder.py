"""
excel_builder.py
Builds the final downloadable workbook with 4 sheets:
  1. Raw Data      - untouched original sheet (zero data loss guarantee)
  2. Cleaned Data  - structured, typed, forward-filled version
  3. Summary       - audit log + rule-based insights + top tables
  4. Dashboard     - native Excel charts (editable/interactive, not images)
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="1F2937")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_df(ws, df, start_row=1, start_col=1, header=True):
    """Write a DataFrame to a worksheet with basic styling, return end row."""
    r = start_row
    if header:
        for j, col_name in enumerate(df.columns):
            cell = ws.cell(row=r, column=start_col + j, value=str(col_name))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
        r += 1
    for _, row in df.iterrows():
        for j, val in enumerate(row):
            if pd.isna(val):
                val = None
            elif hasattr(val, "to_pydatetime"):
                val = val.to_pydatetime()
            cell = ws.cell(row=r, column=start_col + j, value=val)
            cell.border = BORDER
        r += 1
    # autosize columns roughly
    for j, col_name in enumerate(df.columns):
        col_letter = get_column_letter(start_col + j)
        max_len = max(
            [len(str(col_name))] + [len(str(v)) for v in df.iloc[:, j].astype(str).head(200)]
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)
    return r


def build_workbook(raw_df, cleaned_df, audit_log, monthly_df, buyers_df,
                    products_df, insights, output_path):
    wb = Workbook()

    # ---------------- Raw Data ----------------
    ws_raw = wb.active
    ws_raw.title = "Raw Data"
    ws_raw.cell(row=1, column=1, value="Original uploaded data (unmodified)").font = TITLE_FONT
    _write_df(ws_raw, raw_df, start_row=3, header=False)

    # ---------------- Cleaned Data ----------------
    ws_clean = wb.create_sheet("Cleaned Data")
    ws_clean.cell(row=1, column=1, value="Cleaned & Structured Data").font = TITLE_FONT
    _write_df(ws_clean, cleaned_df, start_row=3)

    # ---------------- Summary ----------------
    ws_sum = wb.create_sheet("Summary")
    ws_sum.cell(row=1, column=1, value="Summary & Insights").font = TITLE_FONT
    r = 3
    ws_sum.cell(row=r, column=1, value="Cleaning steps performed:").font = Font(bold=True)
    r += 1
    for line in audit_log:
        ws_sum.cell(row=r, column=1, value=f"• {line}")
        ws_sum.column_dimensions["A"].width = 110
        r += 1
    r += 1
    ws_sum.cell(row=r, column=1, value="Key insights & suggestions:").font = Font(bold=True)
    r += 1
    for line in insights:
        ws_sum.cell(row=r, column=1, value=f"• {line}")
        r += 1
    r += 2

    ws_sum.cell(row=r, column=1, value="Monthly Sales").font = Font(bold=True)
    r += 1
    monthly_start = r
    if not monthly_df.empty:
        r = _write_df(ws_sum, monthly_df, start_row=r)
    r += 2

    ws_sum.cell(row=r, column=1, value="Top Buyers").font = Font(bold=True)
    r += 1
    buyers_start = r
    if not buyers_df.empty:
        r = _write_df(ws_sum, buyers_df, start_row=r)
    r += 2

    ws_sum.cell(row=r, column=1, value="Top Products").font = Font(bold=True)
    r += 1
    products_start = r
    if not products_df.empty:
        r = _write_df(ws_sum, products_df, start_row=r)

    # ---------------- Dashboard ----------------
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash.cell(row=1, column=1, value="Sales Dashboard").font = TITLE_FONT
    ws_dash.sheet_view.showGridLines = False

    anchor_row = 3
    if not monthly_df.empty:
        n = len(monthly_df)
        line = LineChart()
        line.title = "Monthly Sales Trend"
        line.y_axis.title = "Sales Value"
        line.x_axis.title = "Month"
        data = Reference(ws_sum, min_col=2, min_row=monthly_start, max_row=monthly_start + n)
        cats = Reference(ws_sum, min_col=1, min_row=monthly_start + 1, max_row=monthly_start + n)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.width, line.height = 24, 11
        ws_dash.add_chart(line, f"A{anchor_row}")

        growth_bar = BarChart()
        growth_bar.title = "Month-over-Month Growth %"
        data2 = Reference(ws_sum, min_col=5, min_row=monthly_start, max_row=monthly_start + n)
        growth_bar.add_data(data2, titles_from_data=True)
        growth_bar.set_categories(cats)
        growth_bar.width, growth_bar.height = 24, 11
        ws_dash.add_chart(growth_bar, f"M{anchor_row}")
        anchor_row += 23

    if not buyers_df.empty:
        n = len(buyers_df)
        bar = BarChart()
        bar.title = "Top Buyers by Sales Value"
        bar.type = "bar"
        data = Reference(ws_sum, min_col=2, min_row=buyers_start, max_row=buyers_start + n)
        cats = Reference(ws_sum, min_col=1, min_row=buyers_start + 1, max_row=buyers_start + n)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.width, bar.height = 24, 12
        ws_dash.add_chart(bar, f"A{anchor_row}")

    if not products_df.empty:
        n = len(products_df)
        pie = PieChart()
        pie.title = "Top Products Share"
        data = Reference(ws_sum, min_col=2, min_row=products_start, max_row=products_start + n)
        cats = Reference(ws_sum, min_col=1, min_row=products_start + 1, max_row=products_start + n)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        pie.width, pie.height = 24, 12
        ws_dash.add_chart(pie, f"M{anchor_row}")

    wb.save(output_path)
    return output_path
